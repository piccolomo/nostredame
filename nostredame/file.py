import os
from nostredame.time import time_class
from nostredame.values import values_class
import openpyxl
from datetime import datetime

# Path Manipulation
def correct_path(folder):
    folder = os.path.normpath(folder)
    folder = folder.replace("~", user_folder)
    is_relative = folder[ : len(os.sep)] != os.sep 
    folder = os.path.join(input_folder, folder) if is_relative else folder
    return folder

def join_paths(*args):
    return os.path.join(*args)

def base_name(path):
    return os.path.basename(correct_path(path))

def create_folder(path):
    os.makedirs(path) if not os.path.exists(path) else None

add_extension = lambda path, extension: path + "." + extension if "." not in base_name(path) else path


# Folders
user_folder = os.path.expanduser("~")
forecast_folder = join_paths(user_folder, "Documents", "Forecast")

code_folder = join_paths(forecast_folder, "code")
input_folder = join_paths(forecast_folder, "input")
output_folder = join_paths(forecast_folder, "output")

create_folder(code_folder)
create_folder(input_folder)
create_folder(output_folder)


# Read Files
def read_table(path, delimiter = ",", header = False):
    path = add_extension(path, 'csv')
    text = read_text(path)
    text = text[header : ]
    data = [el.replace("\n", "").split(delimiter) for el in text]
    return data

def read_text(path):
    path = correct_path(path)
    with open(path, "r") as file:
        text = file.readlines()
    return text

def write_text(path, text):
    file = open(path, "w")
    file.write(text)
    file.close()


def read_excel_data(path, log = True):
    path = add_extension(path, 'xlsx')
    path = correct_path(path)
    print(path)
    print("reading excel file in", path) if log else None
    workbook = openpyxl.load_workbook(path)
    names = workbook.sheetnames
    sheets = [workbook[name] for name in names]
    matrices = {}
    for sheet in sheets:
        matrix = []
        for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row, min_col = 1, max_col = sheet.max_column):
            line = []
            for cell in row:
                el = cell.value
                el = el.strftime('%d/%m/%Y') if isinstance(el, datetime) else str(el)
                line.append(el)
            matrix.append(line)
        matrices[sheet.title.strip()] = matrix
    workbook.close()
    print("excel file read") if log else None
    return matrices
